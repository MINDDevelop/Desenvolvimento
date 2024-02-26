import Functions as tt 
import cacluldora_BS as BS 
from datetime import datetime,timedelta
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import requests

caminho=r'C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Arquivos'














#################################################################################################################################################
# df1=pd.DataFrame(pd.read_excel(r'C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Inserir_planilha_atm.xlsx'))
# df2=pd.DataFrame(pd.read_excel(r'C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Inserir_planilha_not_atm.xlsx'))
# df2['VI_ask']=np.nan
# df2['VI_bid']=np.nan
# df_concatenado = pd.concat([df1, df2],ignore_index=True)
# df_concatenado['days_to_maturity']=df_concatenado.apply(BS.calcular_du,axis=1)
# df_concatenado['pricer_ask']=df_concatenado.apply(BS.calcular_bs_ask,axis=1)
# df_concatenado['pricer_bid']=df_concatenado.apply(BS.calcular_bs_bid,axis=1)
# df_concatenado=df_concatenado[['ativo_alvo','symbol','category','strike','due_date','bid','ask','close','tmoney','VI_bid','VI_ask','days_to_maturity','pricer_bid','pricer_ask']]
# df_concatenado.to_excel('Base_dados.xlsx')



# caminho_excel1 = r"C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\rel_final.xlsx"
# caminho_excel2 = r"C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Base_dados.xlsx"
# caminho_saida = 'final.xlsx'

# # Carregar os arquivos Excel
# workbook1 = load_workbook(caminho_excel1)
# workbook2 = load_workbook(caminho_excel2)

# # Obter a primeira planilha do segundo arquivo Excel
# sheet2 = workbook2.active

# # Adicionar a planilha do segundo arquivo Excel como uma nova sheet ao primeiro arquivo Excel
# workbook1.create_sheet(title='Base de dados', index=0)
# nova_sheet = workbook1['Base de dados']

# for row in sheet2.iter_rows():
#     for cell in row:
#         nova_sheet[cell.coordinate].value = cell.value

# # Salvar as alterações no primeiro arquivo Excel
# workbook1.save(caminho_saida)

# print(f"As duas planilhas foram combinadas e salvas em '{caminho_saida}'.")

#####################################################################################
                        #salvar o banco de dados na planilha principal
#####################################################################################
# caminho_arquivo_origem = r"C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Vols_final_mes.xlsx"
# caminho_arquivo_destino = r"C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\final.xlsx"

# nome_planilha_origem = 'Sheet1'  # Planilha de onde os dados serão copiados
# nome_planilha_destino = 'Volatilidade (06.02)' # Planilha onde os dados serão colados

# # Carregar os arquivos Excel
# wb_origem = load_workbook(caminho_arquivo_origem)
# wb_destino = load_workbook(caminho_arquivo_destino)

# # Verificar se as planilhas existem nos arquivos
# if nome_planilha_origem in wb_origem.sheetnames and nome_planilha_destino in wb_destino.sheetnames:
#     # Obter as planilhas
#     planilha_origem = wb_origem[nome_planilha_origem]
#     planilha_destino = wb_destino[nome_planilha_destino]

#     # Copiar dados da planilha de origem para a planilha de destino
#     for row_index, row in enumerate(planilha_origem.iter_rows(), start=1):
#         for col_index, cell in enumerate(row, start=1):
#             # Obter o valor da célula na planilha de origem
#             valor_celula_origem = cell.value
#             # Obter a célula correspondente na planilha de destino
#             celula_destino = planilha_destino.cell(row=row_index, column=col_index)
#             # Definir o valor da célula na planilha de destino
#             celula_destino.value = valor_celula_origem

#     # Salvar as alterações no arquivo de destino
#     wb_destino.save(caminho_arquivo_destino)
#     print(f'Dados da planilha "{nome_planilha_origem}" do arquivo "{caminho_arquivo_origem}" copiados para a planilha "{nome_planilha_destino}" do arquivo "{caminho_arquivo_destino}" com sucesso.')
# else:
#     print('Uma ou ambas as planilhas não foram encontradas nos arquivos.')
def get_token(email,senha): #OK
    
    body = {"email": email,"password": senha}
    
    ## CHAMADA NA API
    r = requests.post('https://api.oplab.com.br/v3/domain/users/authenticate',json=body).json()['access-token']
    return r

email='victor.ferrando@gmail.com'
senha='899513Vi!'
Token=get_token(email,senha)

print(get_token(email,senha))
###################################
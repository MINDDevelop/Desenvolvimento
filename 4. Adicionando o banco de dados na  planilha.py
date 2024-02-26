import Functions as tt 
import cacluldora_BS as BS 
from datetime import datetime,timedelta
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
caminho=r'C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Arquivos'


caminho_arquivo_origem = r"C:\Users\vgonçalves\Desktop\Desenvolvimento\Desenvolvimento\Arquivos\Base_de_dados.xlsx"
caminho_arquivo_destino = r"C:\Users\vgonçalves\OneDrive\Documents\Planilha final.xlsx"

nome_planilha_origem = 'Sheet1'  # Planilha de onde os dados serão copiados                  
nome_planilha_destino = 'Base de dados' # Planilha onde os dados serão colados

# Carregar os arquivos Excel
wb_origem = load_workbook(caminho_arquivo_origem)
wb_destino = load_workbook(caminho_arquivo_destino)

# Verificar se as planilhas existem nos arquivos
if nome_planilha_origem in wb_origem.sheetnames and nome_planilha_destino in wb_destino.sheetnames:
    # Obter as planilhas
    planilha_origem = wb_origem[nome_planilha_origem]
    planilha_destino = wb_destino[nome_planilha_destino]

    # Copiar dados da planilha de origem para a planilha de destino
    for row_index, row in enumerate(planilha_origem.iter_rows(), start=1):
        for col_index, cell in enumerate(row, start=1):
            # Obter o valor da célula na planilha de origem
            valor_celula_origem = cell.value
            # Obter a célula correspondente na planilha de destino
            celula_destino = planilha_destino.cell(row=row_index, column=col_index)
            # Definir o valor da célula na planilha de destino
            celula_destino.value = valor_celula_origem

    # Salvar as alterações no arquivo de destino
    wb_destino.save(caminho_arquivo_destino)
    print(f'Dados da planilha "{nome_planilha_origem}" do arquivo "{caminho_arquivo_origem}" copiados para a planilha "{nome_planilha_destino}" do arquivo "{caminho_arquivo_destino}" com sucesso.')
else:
    print('Uma ou ambas as planilhas não foram encontradas nos arquivos.')